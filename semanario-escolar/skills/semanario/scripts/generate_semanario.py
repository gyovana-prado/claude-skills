#!/usr/bin/env python3
"""
Gerador de Semanário Escolar - 3º Ano A
Gera a planilha Excel do semanário no formato exato exigido pela escola Adventista.

Uso:
    python generate_semanario.py <input.json> <output.xlsx>
    python generate_semanario.py <input.json> --append-to <existing.xlsx> <output.xlsx>

Modos:
    1. Novo arquivo: cria um .xlsx novo com uma aba para a quinzena
    2. Append: abre um .xlsx existente, adiciona uma nova aba com a quinzena,
       e salva no output (pode ser o mesmo arquivo ou um novo)

O input.json deve seguir o schema definido no SKILL.md.
"""

import json
import sys
import shutil
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ── Color scheme matching the school template ──────────────────────────
COLORS = {
    "header_blue": "4472C4",
    "header_dark": "2F5496",
    "seg_blue": "BDD7EE",       # Segunda - light blue
    "ter_purple": "D5A6E6",     # Terça - purple
    "qua_red": "FF9999",        # Quarta - red/salmon
    "qui_orange": "FFD699",     # Quinta - orange
    "sex_green": "C6EFCE",      # Sexta - green
    "intervalo_gray": "D9D9D9",
    "white": "FFFFFF",
    "black": "000000",
}

DAY_COLORS = {
    "SEGUNDA-FEIRA": "BDD7EE",
    "TERÇA-FEIRA": "D5A6E6",
    "QUARTA-FEIRA": "FF9999",
    "QUINTA-FEIRA": "FFD699",
    "SEXTA-FEIRA": "C6EFCE",
}

# Component-specific colors (matching the original template)
COMPONENT_COLORS = {
    "LING. PORT.": "BDD7EE",
    "HISTÓRIA": "E2CCFF",
    "CIÊNCIAS": "FFFF99",
    "ED. FÍSICA": "99FF99",
    "MATEMÁTICA": "FFB3B3",
    "ENSINO REL.": "FFD699",
    "BILÍNGUE": "B3D9FF",
    "ARTE": "FF99CC",
    "CULT. MAKER": "FF8080",
    "CULT. GERAL": "FF8080",
    "GEOGRAFIA": "99CCFF",
}

# ── Styles ─────────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
subheader_font = Font(name="Arial", bold=True, size=8)
cell_font = Font(name="Arial", size=8)
cell_font_bold = Font(name="Arial", bold=True, size=8)
title_font = Font(name="Arial", bold=True, size=12, color="2F5496")
wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Column structure ───────────────────────────────────────────────────
COLUMNS = {
    "A": ("Aula ", 6),
    "B": ("COMPONENTE CURRICULAR", 18),
    "C": ("Objeto do Conhecimento", 28),
    "D": ("EVIDÊNCIAS DE COMPETÊNCIAS", 35),
    "E": ("DESENVOLVIMENTO( INÍCIO/DESENVOL./CONCLUSÃO)", 45),
    "F": ("RECURSOS", 15),
    "G": ("AVALIAÇÃO", 15),
    "H": ("CPB", 5),
    "I": ("BNCC", 5),
    "J": ("REFERENCIAL", 10),
    "K": ("OUTROS", 8),
}


def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def write_header_row(ws, row, day_label, date_str, day_color):
    """Write the day header row (e.g., 'SEGUNDA-FEIRA- 13/04')"""
    day_text = f"{day_label} - {date_str}" if date_str else day_label
    ws.cell(row=row, column=1, value=day_text)
    fill = PatternFill("solid", fgColor=day_color)
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        if col == 1:
            cell.value = day_text
        apply_cell_style(cell, font=cell_font_bold, fill=fill, border=thin_border)

    # REF. BIBLIOGRÁFICAS header
    ws.cell(row=row, column=8, value="REF. BIBLIOGRÁFICAS")
    for col in range(8, 12):
        cell = ws.cell(row=row, column=col)
        if col == 8:
            cell.value = "REF. BIBLIOGRÁFICAS"
        apply_cell_style(cell, font=cell_font_bold, fill=fill, border=thin_border)


def write_column_headers(ws, row):
    """Write the column header row for aulas"""
    for col_letter, (header_text, _) in COLUMNS.items():
        col_idx = ord(col_letter) - ord("A") + 1
        cell = ws.cell(row=row, column=col_idx, value=header_text)
        apply_cell_style(
            cell,
            font=Font(name="Arial", bold=True, size=7, color="FFFFFF"),
            fill=header_fill,
            alignment=center_alignment,
            border=thin_border,
        )


def write_aula_row(ws, row, aula_data, is_special=False):
    """Write a single aula row"""
    comp = aula_data.get("componente", "")
    comp_color = COMPONENT_COLORS.get(comp, "FFFFFF")
    comp_fill = PatternFill("solid", fgColor=comp_color)

    # Column A: Aula number
    cell = ws.cell(row=row, column=1, value=aula_data.get("numero", ""))
    apply_cell_style(cell, font=cell_font_bold, alignment=center_alignment, border=thin_border)

    # Column B: Componente Curricular
    cell = ws.cell(row=row, column=2, value=comp)
    apply_cell_style(cell, font=cell_font_bold, fill=comp_fill, alignment=wrap_alignment, border=thin_border)

    if is_special:
        # For special classes, only fill the component name and maybe a note
        note = aula_data.get("objeto_conhecimento", "")
        if note:
            cell = ws.cell(row=row, column=3, value=note)
            apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)
        # Still apply borders to remaining cells
        for col in range(4, 12):
            cell = ws.cell(row=row, column=col)
            apply_cell_style(cell, border=thin_border)
        return

    # Column C: Objeto do Conhecimento
    cell = ws.cell(row=row, column=3, value=aula_data.get("objeto_conhecimento", ""))
    apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)

    # Column D: Evidências de Competências
    cell = ws.cell(row=row, column=4, value=aula_data.get("competencias", ""))
    apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)

    # Column E: Desenvolvimento
    cell = ws.cell(row=row, column=5, value=aula_data.get("desenvolvimento", ""))
    apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)

    # Column F: Recursos
    cell = ws.cell(row=row, column=6, value=aula_data.get("recursos", ""))
    apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)

    # Column G: Avaliação
    cell = ws.cell(row=row, column=7, value=aula_data.get("avaliacao", ""))
    apply_cell_style(cell, font=cell_font, alignment=wrap_alignment, border=thin_border)

    # Columns H-K: Reference marks
    for col, key in [(8, "cpb"), (9, "bncc"), (10, "referencial"), (11, "outros")]:
        val = aula_data.get(key, "")
        cell = ws.cell(row=row, column=col, value=val)
        apply_cell_style(cell, font=cell_font, alignment=center_alignment, border=thin_border)


def write_tarefa_row(ws, row, tarefa_text):
    """Write the homework row"""
    text = f"Tarefa de Casa: {tarefa_text}" if tarefa_text else "Tarefa de Casa:"
    cell = ws.cell(row=row, column=1, value=text)
    apply_cell_style(
        cell,
        font=Font(name="Arial", bold=True, italic=True, size=8),
        alignment=wrap_alignment,
        border=thin_border,
    )
    # Merge across the row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)


def write_week_header(ws, row, data):
    """Write the week header block (Professor, Turma, Quinzena, Materiais)."""
    # Row: Header info labels
    ws.cell(row=row, column=1, value="Professor(a)")
    ws.cell(row=row, column=4, value="Turma")
    ws.cell(row=row, column=5, value="QUINZENA ")
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        apply_cell_style(cell, font=subheader_font, border=thin_border)

    # Row+1: Header info values
    ws.cell(row=row + 1, column=1, value=data.get("professora", ""))
    ws.cell(row=row + 1, column=4, value=data.get("turma", ""))
    ws.cell(row=row + 1, column=5, value=data.get("quinzena", ""))
    for col in range(1, 12):
        cell = ws.cell(row=row + 1, column=col)
        apply_cell_style(cell, font=cell_font_bold, border=thin_border)

    # Row+2: Solicitação de Materiais
    materiais = data.get("materiais", "")
    ws.cell(row=row + 2, column=1, value=f"Solicitação de Materiais: {materiais}" if materiais else "Solicitação de Materiais:")
    apply_cell_style(ws.cell(row=row + 2, column=1), font=cell_font_bold, border=thin_border)

    return row + 3  # Return next available row


def _build_sheet_name(data):
    """Build a short, unique sheet name from the quinzena data.

    Examples:
        "13 a 24 de abril de 2025" → "13-24 Abr"
        "28/04 a 09/05"           → "28Abr-09Mai"

    Falls back to the raw quinzena string (truncated to 31 chars, Excel limit).
    """
    quinzena = data.get("quinzena", "Semanário")

    # Try to build a compact name from the semanas dates
    semanas = data.get("semanas", [])
    if semanas:
        all_dias = []
        for sem in semanas:
            all_dias.extend(sem.get("dias", []))
        if all_dias:
            first_date = all_dias[0].get("data", "")   # e.g. "13/04"
            last_date = all_dias[-1].get("data", "")    # e.g. "24/04"

            MESES_ABREV = {
                "01": "Jan", "02": "Fev", "03": "Mar", "04": "Abr",
                "05": "Mai", "06": "Jun", "07": "Jul", "08": "Ago",
                "09": "Set", "10": "Out", "11": "Nov", "12": "Dez",
            }

            try:
                d1, m1 = first_date.split("/")
                d2, m2 = last_date.split("/")
                mes1 = MESES_ABREV.get(m1, m1)
                mes2 = MESES_ABREV.get(m2, m2)

                if m1 == m2:
                    return f"{d1}-{d2} {mes1}"
                else:
                    return f"{d1}{mes1}-{d2}{mes2}"
            except (ValueError, IndexError):
                pass

    # Fallback: truncate quinzena to Excel's 31-char limit
    return quinzena[:31]


def _populate_sheet(ws, data):
    """Populate a worksheet with the semanário content.

    This is the core rendering logic, extracted so it can be used
    both for new workbooks and for appended sheets.
    """
    special_classes = ["ED. FÍSICA", "ARTE", "BILÍNGUE", "CULT. MAKER", "CULT. GERAL", "BILIÍNGUE"]

    # Set column widths
    for col_letter, (_, width) in COLUMNS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row 1: Empty spacer ──
    current_row = 1

    # ── Write each week with its own header block ──
    semanas = data.get("semanas", [])
    for sem_idx, semana in enumerate(semanas):
        current_row = write_week_header(ws, current_row + 1, data)

        for dia in semana.get("dias", []):
            dia_semana = dia.get("dia_semana", "")
            data_str = dia.get("data", "")
            day_color = DAY_COLORS.get(dia_semana, "FFFFFF")

            write_header_row(ws, current_row, dia_semana, data_str, day_color)
            current_row += 1

            write_column_headers(ws, current_row)
            current_row += 1

            for aula in dia.get("aulas", []):
                comp = aula.get("componente", "")
                is_special = comp.upper() in [s.upper() for s in special_classes]
                write_aula_row(ws, current_row, aula, is_special)
                max_content_len = max(
                    len(str(aula.get("desenvolvimento", ""))),
                    len(str(aula.get("competencias", ""))),
                    len(str(aula.get("objeto_conhecimento", ""))),
                    1,
                )
                row_height = min(max(20, max_content_len // 3), 120)
                ws.row_dimensions[current_row].height = row_height
                current_row += 1

            write_tarefa_row(ws, current_row, dia.get("tarefa_casa", ""))
            current_row += 1


def generate_semanario(data, output_path, append_to=None):
    """Generate the semanário Excel file from structured data.

    Args:
        data: dict with the semanário content (semanas, aulas, etc.)
        output_path: where to save the final .xlsx
        append_to: optional path to an existing .xlsx to append a new sheet to.
                   If provided, opens that file and adds a new tab.
                   If None, creates a brand new workbook.
    """
    sheet_name = _build_sheet_name(data)

    if append_to and Path(append_to).exists():
        # ── APPEND MODE: open existing workbook, add new sheet ──
        # If output is the same file, work in-place. Otherwise, copy first.
        append_path = Path(append_to)
        output_p = Path(output_path)

        if append_path.resolve() != output_p.resolve():
            shutil.copy2(append_to, output_path)
            wb = load_workbook(output_path)
        else:
            wb = load_workbook(append_to)

        # Ensure unique sheet name
        existing_names = wb.sheetnames
        base_name = sheet_name
        counter = 2
        while sheet_name in existing_names:
            sheet_name = f"{base_name} ({counter})"
            counter += 1

        ws = wb.create_sheet(title=sheet_name)
        _populate_sheet(ws, data)

        wb.save(output_path)
        print(f"Nova aba '{sheet_name}' adicionada ao semanário: {output_path}")
        print(f"Abas no arquivo: {wb.sheetnames}")
    else:
        # ── NEW FILE MODE: create from scratch ──
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        _populate_sheet(ws, data)

        wb.save(output_path)
        print(f"Semanário gerado com sucesso: {output_path}")

    return output_path


def main():
    append_to = None
    args = sys.argv[1:]

    # Parse --append-to flag
    if "--append-to" in args:
        idx = args.index("--append-to")
        if idx + 1 >= len(args):
            print("Erro: --append-to precisa de um caminho para o arquivo existente.")
            sys.exit(1)
        append_to = args[idx + 1]
        args = args[:idx] + args[idx + 2:]

    if len(args) < 2:
        print("Uso:")
        print("  python generate_semanario.py <input.json> <output.xlsx>")
        print("  python generate_semanario.py <input.json> --append-to <existing.xlsx> <output.xlsx>")
        print()
        print("Exemplos:")
        print("  # Criar novo arquivo:")
        print("  python generate_semanario.py quinzena1.json semanario_2bim.xlsx")
        print()
        print("  # Adicionar nova aba em arquivo existente:")
        print("  python generate_semanario.py quinzena2.json --append-to semanario_2bim.xlsx semanario_2bim.xlsx")
        sys.exit(1)

    input_path = args[0]
    output_path = args[1]

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    generate_semanario(data, output_path, append_to=append_to)


if __name__ == "__main__":
    main()
